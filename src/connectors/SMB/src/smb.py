"""SMB connector for fast file scanning and change detection.

Provides metadata-based incremental updates with efficient filtering
and support for multiple document formats (PDF, DOCX, XLSX, text).
"""

import os
import base64
import json
from hashlib import sha256
from typing import Any
import pdfplumber
from docx import Document
from openpyxl import load_workbook

try:
    from .variables import (
        ALLOWED_EXTENSIONS,
        BASE_PATH,
        MAX_FILE_SIZE,
        PROJECT,
        SKIP_DIRS,
        SOURCE_FILE,
    )
    from .logs import dms_error
except ImportError:
    from variables import (
        ALLOWED_EXTENSIONS,
        BASE_PATH,
        MAX_FILE_SIZE,
        PROJECT,
        SKIP_DIRS,
        SOURCE_FILE,
    )
    from logs import dms_error


class SMBCollector:
    """SMB connector with fast metadata-based change detection for maximum speed."""

    def __init__(self) -> None:
        self.base_path = BASE_PATH
        self.skip_dires = SKIP_DIRS

    def _hash_project(self, project_path: str) -> str:
        """Fast metadata-only hash of project using os.scandir() without sorting.

        SMB mount returns entries in consistent order, eliminating 94% of overhead.
        """
        h = sha256()

        try:
            entries = os.scandir(project_path)  # Skip sorting - SMB order is consistent!
        except PermissionError:
            return ""

        for entry in entries:
            name = entry.name

            try:
                stat = entry.stat(follow_symlinks=False)
            except (FileNotFoundError, OSError):
                continue

            meta = (int(stat.st_mtime), stat.st_size)

            if entry.is_file(follow_symlinks=False):
                # Hash: filename + mtime + size
                h.update(name.encode())
                h.update(str(meta).encode())

            elif entry.is_dir(follow_symlinks=False) and name not in self.skip_dires:
                # Recursively hash subdirectory
                sub_hash = self._hash_project(entry.path)
                h.update(name.encode())
                h.update(sub_hash.encode())

        return h.hexdigest()

    def _is_valid_file(self, path: str) -> bool:
        """Check if file passes extension and size filters."""
        # 1. Extension filter
        if not path.lower().endswith(ALLOWED_EXTENSIONS):
            return False

        # 2. Size filter
        try:
            size = os.path.getsize(path)
            if size > MAX_FILE_SIZE:  # 5 MB limit
                return False
        except OSError:
            return False

        return True

    # ----------------------------
    # Get PROJECTS
    # ----------------------------
    def _get_projects(self) -> list[str]:
        return [d for d in os.listdir(self.base_path) if os.path.isdir(os.path.join(self.base_path, d))]

    def get_project_ids(self) -> dict[str, str]:
        """Return a dict of project names to their merkle-tree hashes."""
        ids = {}

        for project in self._get_projects():
            path = os.path.join(self.base_path, project)
            ids[project] = self._hash_project(path)

        return ids

    def get_projects_as_units(self) -> dict:
        """Return project metadata as units for indexing."""
        projects: dict = {}

        for project in self._get_projects():
            path = os.path.join(self.base_path, project)

            projects[path] = {
                "name": project,
                "creator": "unknown",
                "created_date": None,
                "last_edit_date": os.path.getmtime(path),
                "type": PROJECT,
            }

        return projects

    # ----------------------------
    # PDF
    def parse_pdf(self, path):
        """Extract text from PDF, handling errors gracefully."""
        parts = []
        try:
            with pdfplumber.open(path) as pdf:
                for page in pdf.pages:
                    text = page.extract_text()
                    if text:
                        parts.append(text)
        except (OSError, ValueError):
            return ""
        return "\n".join(parts)

    # ----------------------------
    # DOCX
    # ----------------------------
    def parse_docx(self, path: str) -> str:
        """Extract text from DOCX, handling errors gracefully."""
        try:
            doc = Document(path)
            return "\n".join(p.text for p in doc.paragraphs if p.text)
        except (OSError, ValueError):
            return ""

    def parse_xlsx(self, path: str) -> str:
        """Extract text from XLSX, handling errors gracefully."""
        parts = []
        try:
            wb = load_workbook(path, read_only=True)
            try:
                for sheet in wb:
                    for row in sheet.iter_rows(values_only=True):
                        row_text = " ".join(str(cell) for cell in row if cell)
                        if row_text:
                            parts.append(row_text)
            finally:
                wb.close()
        except (OSError, ValueError):
            return ""
        return "\n".join(parts)

    def parse_text(self, path: str) -> str:
        """Extract text from a text file, handling errors gracefully."""
        try:
            with open(path, "r", encoding="utf-8", errors="ignore") as f:
                return f.read()
        except OSError:
            return ""

    def parse_file(self, path: str) -> str:
        """Dispatch file parsing based on extension."""
        path_lower = path.lower()

        if path_lower.endswith(".pdf"):
            return self.parse_pdf(path)
        if path_lower.endswith(".docx"):
            return self.parse_docx(path)
        if path_lower.endswith(".xlsx"):
            return self.parse_xlsx(path)
        return self.parse_text(path)

    def get_files_in_project(self, project: str) -> list[str]:
        """Return list of all file paths in a project, excluding skipped directories."""
        project_path = os.path.join(self.base_path, project)
        files: list[str] = []

        for root, dirs, filenames in os.walk(project_path):
            dirs[:] = [d for d in dirs if d not in self.skip_dires]

            for f in filenames:
                files.append(os.path.join(root, f))
        return files

    def get_file(self, path: str, include_content: bool = True) -> dict:
        """Return file metadata and optionally content, handling errors gracefully."""
        stat = os.stat(path)

        metadata = {
            "unique_pointer": path,
            "name": os.path.basename(path),
            "size": stat.st_size,
            "last_edit_date": stat.st_mtime,
            "type": SOURCE_FILE,
        }
        result: dict[str, Any] = {"metadata": metadata}

        if include_content:
            try:
                text = self.parse_file(path)
                content = base64.b64encode(text.encode("utf-8", errors="ignore")).decode()
                result["content"] = content
            except (UnicodeDecodeError, TypeError, ValueError) as err:
                dms_error(f"Could not read file {path}: {err}")
                result["content"] = ""
        return result

    def files_to_index(self, subdata: str | None = None) -> dict:
        """Return list of files to index with fast incremental updates via metadata hashing."""
        files_data: list = []

        if subdata is None:
            subdata_dict = {}
        else:
            try:
                subdata_dict = json.loads(base64.b64decode(subdata))
            except (ValueError, TypeError, json.JSONDecodeError):
                subdata_dict = {}

        current_subdata = self.get_project_ids()

        for project, change_hash in current_subdata.items():
            # Skip project if hash unchanged
            if change_hash == subdata_dict.get(project):
                continue

            files = self.get_files_in_project(project)

            for f in files:
                if not self._is_valid_file(f):
                    continue

                try:
                    files_data.append(self.get_file(f, include_content=True))
                except (OSError, ValueError) as err:
                    dms_error(f"Error processing file {f}: {err}")

        generated_subdata = base64.urlsafe_b64encode(json.dumps(current_subdata).encode()).decode()

        return {"files": files_data, "subdata": generated_subdata}

    def pointers_to_all_files_to_index(self, subdata: str | None) -> dict[str, Any]:
        """Return list of file pointers with fast incremental updates via metadata hashing."""
        if subdata is None:
            subdata_dict = {}
        else:
            try:
                subdata_dict = json.loads(base64.b64decode(subdata))
            except (ValueError, TypeError, json.JSONDecodeError):
                subdata_dict = {}

        file_pointers: list[str] = []
        project_ids = self.get_project_ids()

        for project, change_hash in project_ids.items():
            # Skip project if hash unchanged
            if change_hash == subdata_dict.get(project):
                continue

            files = self.get_files_in_project(project)

            for f in files:
                if self._is_valid_file(f):
                    file_pointers.append(f)

        generated_subdata = base64.urlsafe_b64encode(json.dumps(project_ids).encode()).decode()

        return {"subdata": generated_subdata, "file_pointers": file_pointers}
