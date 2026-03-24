import os
import base64
import json
from hashlib import md5
from typing import Any
from concurrent.futures import ThreadPoolExecutor, as_completed
import pdfplumber
from docx import Document
from openpyxl import load_workbook

from variables import PROJECT, SOURCE_FILE
from logs import dms_error
import variables

class SMBCollector:
    """SMB connector methods as GitLab, at least as close as possible"""

    def __init__(self) -> None:
        self.base_path = variables.BASE_PATH
        self.skip_dires = variables.SKIP_DIRS

    def _hash_project(self, project_path: str) -> str:
        hash_md5 = md5()

        for root, _, files in os.walk(project_path):
            for f in sorted(files):  # important: deterministic order
                file_path = os.path.join(root, f)
                try:
                    stat = os.stat(file_path)
                    hash_md5.update(str(stat.st_mtime).encode())
                    hash_md5.update(str(stat.st_size).encode())
                except (OSError, ValueError):
                    continue

        return hash_md5.hexdigest()

    def _is_valid_file(self, path: str) -> bool:
        # 1. Extension filter
        if not path.lower().endswith(variables.ALLOWED_EXTENSIONS):
            return False

        # 2. Size filter
        try:
            size = os.path.getsize(path)
            if size > variables.MAX_FILE_SIZE:  # 5 MB limit
                return False
        except OSError:
            return False

        return True


    def _get_file_signature(self, path: str) -> str:
        try:
            stat = os.stat(path)
            raw = f"{path}-{stat.st_size}-{stat.st_mtime}"
            return md5(raw.encode()).hexdigest()
        except OSError:
            return ""

    def _build_snapshot(self) -> dict[str, str]:
        snapshot = {}

        for project in self._get_projects():
            files = self.get_files_in_project(project)

            for f in files:
                if not self._is_valid_file(f):
                    continue

                try:
                    stat = os.stat(f)
                    raw = f"{stat.st_size}-{stat.st_mtime}"
                    snapshot[f] = md5(raw.encode()).hexdigest()
                except OSError:
                    continue

        return snapshot

    # ----------------------------
    # Get PROJECTS
    # ----------------------------
    def _get_projects(self) -> list[str]:
        return [
            d for d in os.listdir(self.base_path)
            if os.path.isdir(os.path.join(self.base_path, d))
        ]

    def get_project_ids(self) -> dict[str, str]:
        """Return a dict of project names to their change hashes."""
        ids = {}

        for project in self._get_projects():
            path = os.path.join(self.base_path, project)
            ids[project] = self._hash_project(path)

        return ids
    def get_projects_as_units(self) -> dict:
        """Return project metadata as units for indexing."""
        projects: dict = {}

        for project in self._get_projects():
            path = os.path.join(self.base_path, project)

            projects[path] = {
                "name": project,
                "creator": "unknown",
                "created_date": None,
                "last_edit_date": os.path.getmtime(path),
                "type": PROJECT,
            }

        return projects
    # ----------------------------
    # PDF
    def parse_pdf(self,path):
        """Extract text from PDF, handling errors gracefully."""
        parts = []
        try:
            with pdfplumber.open(path) as pdf:
                for page in pdf.pages:
                    text = page.extract_text()
                    if text:
                        parts.append(text)
        except (OSError, ValueError):
            return ""
        return "\n".join(parts)
    # ----------------------------
    # DOCX
    # ----------------------------
    def parse_docx(self,path):
        """Extract text from DOCX, handling errors gracefully."""
        try:
            doc = Document(path)
            return "\n".join(p.text for p in doc.paragraphs if p.text)
        except (OSError, ValueError):
            return ""

    def parse_xlsx(self, path):
        """Extract text from XLSX, handling errors gracefully."""
        parts = []
        try:
            wb = load_workbook(path, read_only=True)
            try:
                for sheet in wb:
                    for row in sheet.iter_rows(values_only=True):
                        row_text = " ".join(str(cell) for cell in row if cell)
                        if row_text:
                            parts.append(row_text)
            finally:
                wb.close()
        except (OSError, ValueError):
            return ""
        return "\n".join(parts)

    def parse_text(self, path):
        """Extract text from a text file, handling errors gracefully."""
        try:
            with open(path, "r", encoding="utf-8", errors="ignore") as f:
                return f.read()
        except OSError:
            return ""

    def parse_file(self, path: str) -> str:
        """Dispatch file parsing based on extension."""
        path_lower = path.lower()

        if path_lower.endswith(".pdf"):
            return self.parse_pdf(path)
        elif path_lower.endswith(".docx"):
            return self.parse_docx(path)
        elif path_lower.endswith(".xlsx"):
            return self.parse_xlsx(path)
        else:
            return self.parse_text(path)

    def get_files_in_project(self, project: str) -> list[str]:
        """Return list of all file paths in a project, excluding skipped directories."""
        project_path = os.path.join(self.base_path, project)
        files: list[str] = []

        for root, dirs, filenames in os.walk(project_path):
            dirs[:] = [d for d in dirs if d not in self.skip_dires]

            for f in filenames:
                files.append(os.path.join(root, f))
        return files



    def get_file(self, path: str, include_content: bool = True) -> dict:
        """Return file metadata and optionally content, handling errors gracefully."""
        stat = os.stat(path)

        metadata = {
            "unique_pointer": path,
            "name": os.path.basename(path),
            "size": stat.st_size,
            "last_edit_date": stat.st_mtime,
            "type": SOURCE_FILE,
        }
        result: dict[str, Any] = {"metadata": metadata}

        if include_content:
            try:
                text = self.parse_file(path)
                content = base64.b64encode(text.encode("utf-8", errors="ignore")).decode()
                result["content"] = content
            except (UnicodeDecodeError, TypeError, ValueError) as err:
                dms_error(f"Could not read file {path}: {err}")
                result["content"]=""
        return result

    def files_to_index(self, subdata: str | None = None) -> dict:
        """Return list of files to index, with optional incremental update based on subdata."""
        files_data: list = []

        if subdata is None:
            subdata_dict = {}
        else:
            subdata_dict = json.loads(base64.b64decode(subdata))
        current_subdata = self.get_project_ids()

        for project, change_hash in current_subdata.items():
            if change_hash == subdata_dict.get(project):
                continue

            files = self.get_files_in_project(project)
            # using threads for Parallel scanning
            with ThreadPoolExecutor(max_workers=8) as executor:
                futures = []

                for f in files:
                    if not self._is_valid_file(f):
                        continue

                    futures.append(executor.submit(self.get_file, f))

                for future in as_completed(futures):
                    try:
                        files_data.append(future.result())
                    except (OSError, ValueError, UnicodeDecodeError, TypeError) as err:
                        dms_error(f"Error processing file: {err}")

        generated_subdata = base64.urlsafe_b64encode(
            json.dumps(current_subdata).encode()
        ).decode()

        return {"files": files_data, "subdata": generated_subdata}

    def pointers_to_all_files_to_index(self, subdata: str | None) -> dict[str, Any]:
        """Return list of file pointers to index, with optional incremental update based on subdata."""
        if subdata is None:
            subdata_dict = {}
        else:
            subdata_dict = json.loads(base64.b64decode(subdata))

        file_pointers: list[str] = []
        project_ids = self.get_project_ids()

        for project, change_hash in project_ids.items():
            if change_hash == subdata_dict.get(project):
                continue

            files = self.get_files_in_project(project)

            for f in files:
                if self._is_valid_file(f):
                    file_pointers.append(f)

        generated_subdata = base64.urlsafe_b64encode(
            json.dumps(project_ids).encode()
        ).decode()

        return {"subdata": generated_subdata, "file_pointers": file_pointers}